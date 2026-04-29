"""
Radiant Homes — SOA Generator Web App
======================================
Upload a P&L workbook, select month, generate all owner statement PDFs.

Run locally:  python app.py
Deploy:       Push to GitHub → connect to Railway.app
"""

import base64
import calendar
import io
import os
import tempfile
import traceback
import zipfile
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file
import openpyxl

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB max upload

# ─── Embedded Logo (dark version for white background) ─────────────────────────
RADIANT_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAhUAAABICAIAAADOCZApAABIPklEQVR42u09d3gU1fbn3pnZ3XQSEkKVIL2JFOmidASkaWiCgAiKiEp9IAr6EBAQsMBDQKUIAgoIRHp91BBEBENAegmhhISE1N2Ze+/vjwPzWzc7k00Bwbfn8/MLu7Nz27mnFyKEAC94wQte8AwEMCIkQQQAEEGAeLfkfxeodwu84AUv5AG4xAQDIAKEIJp3P7z8wwte8IIXPOQfDolKggnKCPESEC//8MIjcSuNgTHmNTN64W8HASCEILJlT+wVkIggGvcSEC//8MIjcRLGIEkSISRXFoLMxstpHnEpgRkD5/xRnDOACgDC7lAdhJBpa460G/7TtHW/EaowpoHQhBAMQAhVcCGE4JypXDDOGedMcHZfCuKcMc696PlPAleqxBgrHFFFCOIE3o02B8bYyZMnHQ4HpW44uqZp5cqVCwsLw111+3NKqf4VkiG3r/KCF/Jxm7kQggtJkqat+23swkM2P5/sjOxP33j2X51qaoxLlBAQKqGyAE6E5HWp/8/yjwcBmqahHO3dbiNGm5mZWbly5fj4+Jx6BqWUc75gwYJBgwZpmibLssvPOeeSJAFAfHx8WlpaaGhoWFgYMhVCiHfbH6mDPn36dEJCgiRJLqdMCGGMlSxZsnLlyvjPR2nmnGmarFhmRh0dNf+Q5BckcU2jPjzt1uwhz773Ym2mZlHZBqABSITwGyn2pPRsRZIEEAABAjgBQSgIEFwtGewX7GsT4GUy/wSQXZTriRMnXr16VZblgvAVi8USHBxcrVq1WrVqVa5c2WKx5JSRveDKyQnB/+ekLCY6HGNMluVVq1bNmTPnjz/+yMjICAoKql+//tixY5s1a4ac27u3j4iKKcvy5MmTly1bZvRM3759ly5dyhhDgeBRmTkHWbF8uen4qPkHZL9AgGwHtSkiU/L3H/71IYmKYR3qqIwrVNU4yJIyec2huT//aQ2UOWMCgAggAgSRJYnaU1P/M7zlG62rMiZkyUsK/in8Qzc3ff/995cvXy6st0uSVKlSpW7duvXp06dKlSqoi7hI0F7wRHR1+6EQQpblUaNGzZw5U/88KSlp8+bNW7du/frrrwcNGsQ597KQRwcsFgs6tDTtL5GvsiwzxlDSeiRQDkAAUM7tAqwSnbMl9t25/6X+oUIwDpyAxgCASJJf4DtfHyaK8nabmg7NClQACImDYFmM+WuM3DNyEAGCS4JyACb4/RG8/OOxB1fKUqRIEVmWLRaLXGCQJIkxdurUqcmTJ9etW/ftt9++fv26LMsuN8cL+ZZnKaVLly6dOXOmoijoY0dA9fHNN988fPgwmr+82/WIgB7jkBMetdgHIpjGuVWi83bEvjNnh+zrL4AxYEj3BRAhOAG77Osz7Ku983bEWWRJqBoAUWWJUAtQCSQCEgEJgFIJBAAQIF5Z5p/MPxhjWiEB2t8lSVIUJTMzc+7cuQ0aNIiKikIW4g0TKqBGIklSVlbWpEmTkENgjC+CpmmSJHHOp02bZqS+eMELZsyDc8a4LMsLd8UN+2If9SnCqSyEIPAXXNJAodxBfQOGfrV3we44q9XGuZA5E0IS4r4WI9AJck/p8MI/mX8UOpljjKmqikLx1atXO3XqNGvWLNTWvbtfkI0lhPz555/nz59H5pFTziWEHD58OC0tLae31gteMAc7F7KsfLfn9JuzdwifAE5lLgQBRlyccyAcxFcWdskiD/n84KKdsZQSqhKALCBerfefDw/JFaELxQAwcuRIxtjo0aMfNT/h48U/AODOnTtCCLcWKlRE7ty5k5iYGBAQYBT46wUv/D/OgCBAuBCcCassfb/v1JuzdgtbCKVZgkvUnQRChcoJOMCmEAdYLINn7gou4iMFW4BTEBTAKyN6+UfhAVq0ZFkeM2ZMREREZGSkl4UUBAIDAwkhbt0byC2CgoKKFi0Kj1g86CPLj420tP+ZHCYugGoatyjSin1nB87YzKxFQBKCESAg3O0AJxIAJ8A1IISCZvMfMHNbeHgY8bF6EwX/F4A+/IuKIvObb755+fJlr3c3n8dGqRCicuXKZcqUcZtbgw/UqVMnKCgI2bZ300wAOYRR/v//zu5xjVkUadWhcwNmbWeWQIkIYCqnigeEggihUUVOUX3/vJRKFFkI77326h/ublqu8aDmkSQYCJ+cnDxmzJhVq1Z5TfP5o3eMMX9//3Hjxg0ZMsRisWBIAtqpkHkAwOjRo717latOLEnSihUrvvjiC5vN5uJJopRqmlaqVKlly5YpivLPZiR2DWyyvCbmwoDp2+1KAJWAaQ5CPHV6U8GoUIHYQJGJcAjiDdP38g93CoQnrm+M/zHiDegLWb16dXR0dMOGDb1WrHwA7vAbb7xx7NixBQsW5DygmTNnPv/883p2uheM8BkALl26dPjwYaNnwsPD4R9tAxQAjDGbLP382/m+07dnU3+FZGsaFVQGIShnwoO1cyIJoIQzTginViK8zg8v//irzAsAJUqUqFu3Lob3uNU8EhMTz5w5k5qaioq/kXmKUqqq6rx58xo2bOhVQfKthQDA/Pnz69at++WXX549e9bhcPj6+taqVWv8+PEdOnTwMmYPwWKxSJKUMzkJ7at+fn641f/AMAQhGCGaZrfK1l+One83eWsWCVBkVTAJKMFoK+HxijlQQrGvlNd45eUfOSg+Y6xZs2YrVqwwfzI+Pn7BggXTpk3DyF237AGN8lFRUTdv3gwPD/cGCOWPf2A+2uDBg/v373/+/PmMjIzg4ODy5csjL/cyjzxp1WgVdPkcd/gfq3gQwRmzypbNv8e/MnlLmgikClEFB2ojoN5P2PD0YhJvmsf/GOTBf470XdM0zrmqqkbNKoQQpUuX/ve//71mzRqr1WoUu4IazJ07d2JiYuB+yVgv5Pn8KEW+brFYqlatWq9ePWQemJ3u3R8vmHOPLE0okrz9j2u9Jm+8ywIlC5G1DIUz8DYW9ELh8g8XmmUSrMI5t9vtHTt2nDhxokn9JXz42LFjUKg50igzuiTSP/zmCijS5szJL3RjHSb565IyLrOwNA+jnfSaHP8BwBj3kaU9sfHdP9mcyq0WRWUimxFZCBt4z9cLHsADiZGglCqKwjl/8803Z82alZiY6NaKhfFCJ06cgMJwTuomCKwEZUINJUlyYWkmNoq8Vp53noYREUcSLMtywVftliHpa8l32XzckJwb5TJ0TkZlEl6Be5Lzc5N6aG6PUq8fpb+2gG1mkPWaP6NpGuKw21EedElQ5/VCQfNRBAguiKRqmkWW98dd7T4pKkX1laygMgIgMUIBKEBhClsEBCdUAgDBsaYi4UJIgvzPlFDUE4xy4u2DG07H6gc31oPCezSqFClSpHHjxuvXr89ZcFSHW7duQYGbHSFXwGt89+7duLi4s2fPJiUlZWZmSpIUFBRUpkyZqlWrPvnkkziQS2+MQmlPorfiwGncunUrLi7uwoULKSkpWVlZFoslJCQkIiKievXqxYsXR0qK0yjIuRa6h0PvPYVBwKdPnz5z5kx8fPzdu3cZY35+fuHh4RUrVqxevbqPj4/z5uuYmldi6uHzyNJkWTY6qXz0O+GcY0FoX19fc2TWF/swQefibpEEWVreEYAwImlMs8rSvj8TXvrkl0QtRLaooGULohAhADgQVrjFcYkQgtwr104I5UJwQjmADH9hIS5s0vlzoyycgnPWnIMWouFXF03wmHLODeWtQmyPpF+TnMTBGZ0edf6hHwz2MjIBVVULyDmwILbdbo+Kivrxxx8PHToUHx+f80kfH5/atWu3a9eud+/e6CTQG1sdP378v//9ryzLznIoxv7XqFGjRYsW5u59rM6CRXCTkpJWr14dFRUVHR2dlJSU8+Hg4OD69et36NChR48exYoVg/z2fMQprV692m0/Ipz8U0899fzzz3sYm4CojJj3+++///DDDzt37jxx4oRbxl+uXLnnnnuuR48ebdu2xULLWEw+NTV1+fLlLuI8mjT9/f179+5ttVpxPvj/tLS05cuXuwyB3yqK8uqrr/r4+OiKDt6xK1eunD17NiEhIT09XVGU8PDwSpUqlStXDoufo3aSa9Kf/k6bzXbnzp1jx44ZqcgAkJqa+tVXX+lZNTnn2bt370IsEqMX+9HJSmpqanp6empqqsPhCAwMLFq0aFBQkM5389QTQQAwTbXKyr4zid0mbbqd7SNbBNUcjMiA7Z5QYShcUkDIvTKKRMrSgBLChF0WVgLCeShz6u+8vUgK3RobnHcvVzprZK5QVbWAqaP6jUAizjlPSkq6fv16SkqKpmk+Pj5hYWHFihULDAzEs8PbVxDijpivL/zGjRtXr169e/duYGAg9pTz9/fHr/IreTxc/oHYcPPmTTB1bxSk5wHuOGNs0aJFs2fPjouL04d2PnvkZFlZWQcPHjx48OCMGTMiIyPHjRtXoUIFjBD78MMPo6Ki3A7Rq1evFi1a6ITVaBqKoiCVmTdvXkJCgsk07ty5s3Xr1q1bt06ZMmXgwIHDhw8vWrRoPvwieJ0+/vjj2NhYo2f69ev3/PPPm09eRz5KqSzLR44cmTp16i+//KLzdRfhCBH94sWLFy9eXLx4cf369ceMGfPSSy/hPmzatGno0KFuh/D39+/WrZvVanWe/+3bt4cMGWI0K6vV2q9fP0T0uLi4lStXbty48cyZM+np6S6PRUREtG7duk+fPg0aNPCEpOI7T5w48f33369cuTI+Ph41Zrf8Iykp6Z133jF5W+vWrQuLf6DEarFYhBAHDx5cu3ZtdHT0pUuXUlNTs7KyGGM2my0kJKR06dJNmzbt3LlzkyZNcKUeNnphjFtk5ci5G90mbbmdYVGsGmhpjMicEAIPykEogAKAEJz62r5YFd2wUtEmTxZjKpdken/VQAhcv3790qVLLgdHCFFV9cknnyxWrJhultGpZFpaWmJiYkZGRnZ2ttVqDQsLK1GihKIoYNywTrcT4BuysrJu3LiRlpZmt9sVRQkJCSlZsiS+AZ/Mh2aga+RJSUl79uyJior67bff4uPjU1NTddHKYrGEhoaWL1++adOmL774Yv369XHh+Yu513/1+++/r1y5ctu2bZcuXbpz5w5+a7PZihUr9vTTT7dt2/all17CfKbCCe537j3AOa9Ro4aRBofLe/nll5HDC1NgjDHGEhMTQ0JCjNwbOPvIyEhPXpgTVFUVQhw7duzZZ5/VZ2hi30DTio6dRYoUmTVrlhAiNTW1aNGiObue2Gw2WZYHDRpkMj00gAghfvnll0qVKjlPw4iUIOfXpxEREbF+/XohxBNPPOF253GXFi5cqC9Z32EhBJIPq9Xq0nwFJ//WW295srf4QHZ29vvvv483hxCi9xQxcXHp+NelS5fLly8LISIjI/XRdcCNLVu27N27d3Vkw/lfunQpICBAlmVFUVx+gs3HhBCXL18eMGCAznhwdOeHnSfWvn37mJgYHQONjmz9+vUdOnTQf2suaeJuuMxQn3NgYODFixf1FZmja//+/d2a7PCTfv364cM//fRTo0aN3M7E5ZO6desuXbrUeQi31xE33aEyIcSvF26W7DMPOi+wdF8EkYtI5CKIXEojv6GRC0nkdw/iPxyFRn4ndV8EnRaV6rvw1wuJQgg7U4XgQnBVtQshZsyYYUR82rVrJ4RwOBy4nkuXLs2ZM6dDhw6lS5f28fHBn6ClulGjRhMnToyNjdVxW3eY6bdVCHH79u0lS5b07NmzXLlyfn5++AZCiL+/f7Vq1d5+++3du3fntquG1EAIcf78+dGjR5csWdKtPzXnOdavX//bb7+12+04olGTmJwj6oOeOnWqZ8+eziwBDTMuYxUrVmz06NG3b9/O09KM4IGEeOKSKKVfffVVcnKyUf1wVFCeeuopyHv8FcrUy5Yte/bZZ/ft24dsAwOEjHyhiEyou8mynJKSMmLEiIEDB27cuDEpKYkx5nA43HYxMWe9kiRNmjSpY8eOZ86cQbaBPzRakR6XhdO4dOlS586dR4wYgRicj30wb8GS6ztRDLly5UqbNm2mTJmC/0TcMtGKMIZbNx6uW7euSZMmUVFRv//+u6Zpdrvd7XyMXNM5weFwMMbOnDnz2muvNWnSZNGiRXa7HU8ZrWEur8VpAMCmTZsaN2788ccf4366IINuNOvXr9/GjRt1i4F5RBnuhqqqRvtcWA7PgICAlJSUl156KTIy8tChQ4ghzp3B9D6hyEEppUePHn311VdbtGhx/PhxWZaRIrjzeRBV1RQZfruc/NLEdQnpfopN0Tj6OQQBTQAV8KCyhQgIACEAuBCKL7+WpnSdGPXH5UQLlVUu+H0bFh6uC/9AeWvLli1bt25VFCUxMXHEiBFPP/3022+/vXHjxvj4+KysLDxlxlhqauqhQ4c+/vjjevXqDR48+Nq1a86eV7RrpaWlTZo0qVatWv369Vu5cuXFixczMjLwDUKI9PT0uLi4OXPmNG/evG3btgcPHvS82QS+3263T5o0qV69ejNmzEDbso634NRADM9RFyVjYmIGDhzYqFGjLVu26BYtD23OkiQtXLiwYcOGK1eu1H2EiNV4hfWxJEm6devWjBkz9FZMBeyjkf/gHCNAPLBYLD/99NPUqVNNKiTiPtauXRvyGH+F5zR79uy+fftmZGQoimLCNowYCd7A7777bsCAAUZJjrnaGbAQ5IQJE/Bs8tQXC6eBhz179uwrV67AQ+/1hMh35syZ5s2b7927V1f884QMyMvj4+M7dep06dKlfGymieC/aNGi+Ph4nTEbEXq9CQo6sT766KPIyEiHw+G2AoIQwt/fH+nyg4iozjf/OH36dMuWLdeuXasoCrpb9Jhv3c3rIuqiJrR79+6mTZsuXboU4x5dVsSBcE21KPLxqyldPvr58l2r5GPhTPtbSlSpTJKt8tVU2nni+hNXkxRKGXNwJETu/Of6J5999tmBAwcaNWo0e/bslJQUVP6cI4ucKXJ2djaSVOQ6yOYVRTlw4EDjxo0nTJhw7do13Vzh8gakDJIkbdu27fnnn58xY4bu4TMnSoqinDx5slmzZhMmTLhz5w4WTEMJL+eh6FYEXQBSFOW333574YUXhg8f7jab1S3aYznzwYMHp6amIuPJOZyzTUiSJIvFcv78+U6dOn355ZcmkU2FzD9wQrjjiN9ugXN++vTpUaNG9erVS9fm3BpAhBChoaH169fPU8wD0ty5c+eOGDECSUD+PPC6koQ6Y/4o79ChQ+fPn68oCp5NPqaBh/231CFGK/DVq1fbtm174cIFlF7zNw3cSXQ8Fm4qDx6x54wZn1QUZfXq1X369NHFPRe2pFu34NEA3PYdO3b89ttvkiR5eBC6Si1JUnp6er9+/T777DP8+V/IB2eKrMRdv9N50o6rycJi04h6V4AkQDKqjkgISJTk+T+JSBI1FQUFgES1TB+bdjFV6frvX/68fkeRFPPF4s3avXt3ixYtzp8/j2wD1cGcnFWnyCjTvPjiiytXrkRWsWrVqpYtW8bGxjrbCZzDwZ15M0pFjLExY8aMHDkSWYjRDFE53rJly7PPPhsTE6PP0HMEQ4Ue1ejPP/+8Q4cOaLkx2Rmc4QcffIAczsNrgoYWpNXvvvvul19+WRAtRM4TihNC9uzZ065dO7clwVHwvHnz5p9//mm3283rXyGWd+3aNSwszHNPDm7Ztm3bhg0bpifNmQuw+sTcSve4qPxZz2bPnv2f//wHpZtc5Wh9Gkbi898i8GZnZ0dGRqLT0sNVuEiFBdxJT7Y6H0tTVdVqta5Zs2bcuHHTp09/XOqA6cHl+dglpAijR4+2WCzvvPOOvmQhhETpwXO3XpywPjkxSwn2VzkBIguQKdgBnEOgCAAnggAFziSelZ0fTzlw8PGRZCEYB0KFayoJkUS2RhUViGQVFy5mNh2xdtO/29UpX8KTNeKiPJSXka1qmta3b98yZco4HI7evXujRJKnN1gsllmzZoWFhY0dO9YtIjHGrFbr+vXru3fv7nA4PLlK5jTWarVu27bthRde2LRpU9GiRd368FGMXrNmzeTJk5EE5enq4UCyLA8fPrxq1aqtW7fO3x2R80pxbty4cePGjVwfNq+/q2teGHvjofEKiWxiYuLAgQN1ScHkKmJEjTPJ1mPpXJS7fGgeR44cGTt2bK693PU8D+eBZFn2sIzxA7VcybL8/vvvHz58WFEUEx0ObQIuKZa6run84UMQ53VTgxEn1sHhcEiSNHPmzA4dOjz33HMu10N3Y+ZM7DKh7G4FpkJMFzDB51x5sx6++d5771WqVAmFPEmSgBB7tn3tnpMNKxez1Ja2/J6kOjghgoBD/DVOVwAQAEqJZtfCi5Bn65diHMBTw7IAIYHgQMWBU7cSUxi1SoJzQRQqOIDQK/gSEJzIwBxWi9SybQU1M33t3hNVS4X4+/rw3OKGPTHp5GSrmqa99NJLiKtuo+zM34CU6oMPPmjatGnTpk1dEAnp+MGDB3v37o1yvTnz0LHXhMxiJFhMTEz37t03btyom+mcn6GU3r59e9iwYSY2WGdnkpEBjXM+dOjQo0eP+vn55SOAsPD7f+hapNmosqxpWv/+/WvXru15hBxejzFjxsTHx5uQPJyhXrAkKCjIz88vKysrNTVVn5W5NporE3U4HEOGDEFZwwgVnGVJQkhwcLDNZsvIyEhNTUUMQ5z4Wwp/IfPYt2/f559/jsY3E/6nGwRsNluRIkWEEGlpaZmZmXpi1MMpZ4Lk3vmq4EEbja7Heo4ePTo6OtoFx/T5F5yyZ2RkPKDl63dN94LqQQFGS8bPBw4ceOzYsdDQULxcikWZ+npzBeDo1Tsbo38ikkLc1SchwKlQGIEgi7r6Xy2bVi2Tv2kfOHm1/cRtd4WFUlXi2QIkQej9zBKhEZvEGSWcOdSPuz9du2wxDuDItoMHWSc57ZC5Si0oaGMWQc7r5skb8FvG2HvvvYfudOeXU0pv3brVq1cvTFU2L77gLG+5lWWd2ZLFYtm1a9fw4cPnzZvnwrTw/s6dO/f69etG6g5eDef55Jweak5nz55dtmzZkCFD8pRIlE/+UXDBGZdRvHjxKVOmeH6H0VgfHR39/fffm2ig+q7Vq1evW7duLVq0KFu2rK+vr6qqCQkJR44c2bFjx7p167KysvLncsBpLFq06OjRoybT0I+qdevWnTt3btq0acmSJS0Wi91uv3z5ckxMTFRU1NatW426lz8EKV7TtPfff9+kvixK6JqmBQcHv/zyy+3bt69Vq1ZwcDDyjz/++GP//v1r1qw5e/asHhr04CasXxJfX9+yZcvabLaEhISbN2/q98otTuqa4o4dO9q0aYPmCFxarVq1sKwOHujt27dRq3a7CovFUrFixZwnhT/39fV1ji0urANC7NLXZbFYMCbNWbc2UusTEhLGjh373Xff3c+QFUJTOZWWbj2hZmTLwVbG3PU8FhKRVXHX/unI55pWLaOp2v9Lrx5IpQQEBwLM3qR6mU/fajx05lYICKOgcqzrfq8XOiGCEaER2WpP4yt2/vF0/xacM0mWkIHlla26WBqMSFZOPuGMNuaCiG7IOnr06Pr16527buNuDx069MqVKyZmK2chDJ3z+lHiQbtNPHI4HIqifP311+3atevcubNO3DFoIj09fdGiRUYCKC7HYrE0bty4WLFi165dO3z4MHr40V6i2w/sdjsAzJs3b9CgQfmx8Xqe/1FYUiS6ejZu3JintA/Ej65du4Jx0Q6cdmho6Pz5801Cm0+ePPniiy9CbsU/8LRee+01fZ5IarOysipVqmSih+Frq1SpsmHDBpMV7d69G2PPzKdhnv/RsGFDt2/Ayb/xxhs5o7xxLVu3bjUZWl9d//79L1y4YLSE9PT0mTNn2mw2cyMkvqpUqVJu8z+wNEiuP69du/bixYsvXLiAy8HMrIEDB+q+Jbe/xYiyHj16OCObc7ggphR89tln4C4tA4cuV65cdna222hD5+gaT9KVjPI/jDD5vffeW79+fXR09MWLFw8dOrR8+fL+/fvjYk3ynPCW/X8eDObPqmq1Ycuhy7dSd/dZGkr3RdBhYZdPfuZCODSVc8YFx/+E4IJzzNW497f+n9MnTGhMMLuqcSG6TFoNHRfKPZbRlxdB5GJ9FBr5DXRfYon8BrosqTnsh3RV5VzLVlUhxMyZn+W6Oc4HbbFYnDm357YXHe0lSXKuTGNC95BqNW/eXA+KRXRav369+ZzxneHh4SNHjty4cWNsbOy5c+diYmIWL17cqVMnnLPRNcRBy5cvj7YTRDMcd//+/UYkCD9s0KDBsWPHdPQ7duxY+/btcdOcN6p8+fIjRozYuXNn/nJBHir/0JP75s2bl6fsFTywU6dOYUicW0TBN1esWPHEiRO4FkxQ0O85Ywyj+PGdI0eONKfdOfkH/nbNmjXmRw4ALVu2xAwdTdOcp4GhHXpqRVpaWseOHXNF3AfBP9q3b29UwwBFJELIzJkzddqnV97Viab+2r179wYFBZkUnygI/8CFDB8+HMPkcsKWLVtCQ0PN71KFChWysrL0oXPuhjn/iIiIKHimlef8A3ne4MGDExIS3L5q3759tWrVMnnPX1J9mcY0LoT4I/6OT+QiiFxMDbL8aOQia5f5h8/fEkKoTMWd4nlbpSYEVxkXQkSfu2HpNo90Xy69/B2JXPTXsRZLkd/Cy99auy88fjVZCJ5tt+fKP5zRtXfv3ps3bz516lRcXNy2bdsGDRqEjMQTwoXvf+aZZ5YsWRIbG3vmzJndu3ePHz8+KChI33wjvuXj44O5ongdHA5HnTp1TMbFCZsc5d69ezEBzvwo8SY6U6Hp06cbYSwhpFixYtevX3e+ufjzF154AR+rVKnSsGHDdu3ahdbXfMND4h96WT1ZlhcsWJDX1Ed8ePLkyUYbjcQrNDT0zJkz6IAyEQn1yL9u3bqZcIKc/EP/lVGVQNy3unXr4qno6bJuAb9NTU2tWbNmrtpMYfEP3JarV6/6+fkZUW28Qp9++ilO0iStmnOenZ0thFi3bp1JDbh88w9c1+uvv47P611n9CBLZCp79uyxWq0mOeSEkOPHj7ukiDvrH5j5bMI/EKOc5QDnhjeFyD9wyZ988okL50bAzEohxK1bt2rWrGni2CeE+Pj4nD17VgjhcKhCiO/3nIIX/iP1WOKWecg9FkGH+S98spEjyykYs2SMa0K8MGk9dFxgi/wWIpeSyG9dRpR6LIb2c1fs+1MIkZ1tF0J85gH/8PPzW7t2rVttPiQkJNdChLi9Q4cOzUl/Tp06VaFCBRNRBn/7ww8/YDEkIcSGDRtMJozPz5gxQ7/vLkeJ9CQlJaV58+ZGhAixOiIiAj12KBYLId566y0wLmTQt29frCjhgn7nzp0bNmzY7t27cf76V+alEx52/rlbD5WmaZUqVUJhIa+OGnQ9/fzzz0YWaswmmT9/fsWKFVVVddHR3G4xSp0BAQFGvXjdeu+Tk5P379/vNvQLfQABAQHLli3z9fVFa6O5mKlpWmBg4PTp0x9a+0W0tEZFRWVkZLitC4Bx1R07dvzXv/6lpzeaHK7ValVVtXPnzhjiUohhsmjbLV26tH4DnQNRkJqgV+C5554bPHiw0eiIG1iULGetVg+3nRhD4SrojLEePXqMHz8eqQwmsunJVZIkYeRIWFjY2rVrUWTOuWrE1aysrLVr1wKAxjkA/H4xEQSTgAi3WV+EAFc7PRNBAESBu88yARJAt4blgTmM3OKEAHA4cvYmAAgP6jaik2Pu3Lldu3ZFWqxrww6H4/nnn0cZy/ziM8Y6duw4Z84cdC85v6FKlSpr1641kavww19//VX/e/HixSbMBl3uo0aNQrkH6/04HyXetaCgoDVr1lSoUMFtJBF+eOnSpV9++QVvBI5o7oTW0+l1pw7m1ZYvX/7LL798/vnnbTabrpeY3/FcDK350ySMqg8ZSQ0lS5YcO3bs4cOHmzdvnlfmgThx/fr1P//8020kDB5V69atu3Xr5uHLEXvKlSvXs2dPvGweRuDExMTcunXLrdMb6dSbb75ZpUoVDM3ycBqtW7du3Ljxw8lRQOQ7cOCA0SVBt9uUKVPA42rYuPARI0YUrhcd2Vv79u2LFClisjkoW/Tt29fIpYxLwFrIj3LbK73lwezZsxHnjZaMLKRChQoffPCBifRDCNmyZQsASBIFgMtJGkiA7cldWswSAMaEJcDarEoYAAgqUyhYWQtKAKBepVCLv2IXkgQON6RGCJCkC0lZOucQuQXdtGrVql+/fhibhFQPCbHFYtE0rVu3bjnja13EER8fHxRH9Dof+htUVa1ZsybaG9y+AZHn6tWrAGC1Wq9du7Zt2za38URIHypXrjx58mTzYsAoRAYHB3/++edGyInXcN26dc7TwLqCbgVESunmzZv/+OMPm82mRy2i915PjdRrAxdQAKJ5pT56yqsOnvCuuXPnTp06NSgoyJNasG4J9+nTp9PS0nLW0NY3dMSIEfiH5xKlEALdDx6WmgGAo0ePGhkNMBhuyJAhHjIkfQ6SJHXo0AEKo4mWh57DkydPGnFiJNk1a9b0nJ/hY/Xq1atcuXL+6pWaQM2aNXPNzKCUli5d2jyAHeNMHmVA3H7ttddKlCiRa/thFCffeuutcuXKub2DSCNiYmKuXbtmkWUBcOduFlALA06Au0j6hBChamWK+pctHpI7LfdkLQQARMUSRcuEBQoVD0W4fS453aHzm1yxH80yJlepT58+5jaM1q1bV6lSxS1uI9EfNGgQRrQbIVJGRgb+ERMTk56e7pYi4WRee+011M5dyrXlLJ5mt9tbt26NyQw5J4Y6xP79+51tBmXKlDEhU8nJya1atZo9e/Yff/yBDExnliZZzPnRmPOK31WrVsVgMuSKmZmZCxYsMKlTgtxv7NixDRs2LFasWD6IC74W2xQaxVCGh4c3atQoT0Xtkfc2aNAgJCQkOTnZQ9n58uXLJvJRgwYNIiIiPOcfOivCIi4POpAXyWtSUhL2RzFab5cuXfKKYSjNNWnS5PTp04UbkRwQEOAJW7XZbAVpBPAoAG5a165dPTFm4rXy9fXt06fPpEmTcu45viQjI+PMmbOlSpVigmdlZwGRBfbhcH0dAOdhAVYfiyQEk+61ICwgsjE/RS4aaD1/5a6wWHM2NMRJZGWx+/wmF7urLMu1a9c28hQiOapTp455TlXjxo1NErYIIZUrVy5XrtyZM2fcMgZnw9H+/fuNKBJjzN/ff/DgwS7xXeYW9ddff33o0KFuG24SQq5cuXLy5EkkFLgQo3QTfP7WrVtoEqhWrdrTTz9dt27d+vXr16lTx7ntW177rRWUf2iaVrt27alTpzp/HhERMWrUKKPwZ2T1f/7558CBA6OiovLdIwXVRiOtv2bNmkFBQfkQfoODg8PCwjzhHzjn69evu6W8+G3Tpk0RezyfBv6wTJkyDyERBGlKQkJCUlKS2/UiSj311FN5RSx8VenSpR/EnAv9yUdT+eCcR0RE1K1b18PNRwLXrVs3rJpsdDUuX74EAIzdSyUzznsGWaYUgAMTQCkUXBcWAKAoMgA3qrIFhHDOuPBoc8LDw7Htmwn1KFq0qM1my8rKMrrOJUuWNDH3ofE2IiLizJkzJo/hH8ePHzfBOovF8tlnn6GamCu5Q+8IWgVMjvL06dP169fHpVWpUqVixYp//vmnW7qhZ3homnby5MmTJ08uX74cAMqVK9egQYPOnTu3adMGLWConeTb8pHn/EEsze2ctzVy5Mjff/992bJlJixEUZRNmzZ98MEHU6ZMUVXV3LHs9sBcWga5fFuqVCnwuIWO8w8VRfH39/eENuGbjaaBgP0E8+GQCAoKCgwMTElJeaCJePjm9PR0txIuDu3r6xsaGpq/9xcpUqTQ5/xwwgr+dsBlVq5c2cfHx0M0xmdq1KhRpUqVkydPuhWEAeDatWv4tyzJAMK4yQloGmcAFCRyr+J6AXeeAoCqaUAoELcJ76i4y9SzvEFfX99cVcxc2w6aG8/xggQGBpogHj6jqqqRKKlbkDBetFAEJpyM3laVc261WkeMGIEqjlu507kaOipnetu3lStXlipVqk+fPm+99dYTTzxh1GjL0zPOK6K7tNDhnH/99ddPP/20CStD9XPq1KmrV6/2pOBgTnA4HCbcHo88f/J4rulvzkdoZEbHiXnCiozuhnkL7kIE3Emj9dpsNpxJPvAp38v3Au42yh+eCxB4rTCBwOi8UlJSAUCWwOZjAcEJcccYBAClt9Pt2SojRGICoIAhWAIIoRmqlpySBVQB7i6uAQC48LFQAOCPm+6YnZ2tO0JMeFWeIFejd0pKirM68uqrrz777LO5xnnqCVsoBONY165dmzZtWr169b766is9l/5h8A+3QpCfn98PP/wQFBRkVE5Ob93++uuvx8XF5aNisHmW2e3bt/NK8pB5cM4RD3I1XuEDRtZMHFpvGJlXcSM1NfXu3bsPh0iZG2Sxz3b+zEE6fnshf5AnvVw/I+xcaYT8Dns2AFAgIQE24IyCEOCmpSvI8tXbaVduJAN6Kgqme3AhAMi5hKQrt9OJRRLutRkCnBf1V3Ti9bfzkDxxbqy/Z/ITLY+QK0lEyQ/uO3usVuuqVauqVaumqqon7Afux1/pzesSExPfeeedPn366B0PHzb/gPt+kapVq37zzTcmxj78KjU1tVevXunp6Z6XDtRbs5l8i81T8+ELcjgcaWlpnk8DI+6NAG0F+YDU1FS0jD0EIz72i3VrMyWE2O32XFvWG+2Pl3/8jbqLsSZw75qWDvEBxgUobp+RZWJPc+w7nQgAlGu8YL0I0aXx69lER4bDQhgDS07/OTrtnwjzBQAioMA866EChvzCw7WvOhvf0FxZokSJnTt3tm/fXvcp5CzWa3RbMWzPYrEsX768b9++WAo2r7eeFtbCNE17+eWXx48fb5KBgV+dOHHijTfeQK7jyXRxLypWrAgGbScAIDY29saNG3lyHuCT169fz5lZZvI8OlrcxkgAwI4dO1RVNerXa/RaIcTZs2cht0JYhUVoypQpEx4eDsa91g8fPpxX/oEoi6vwwqMJtcuFANEYEPfebCGAKutiLorCoIno0lh3+AJQq5HzXAgACg0qhgMAEAr3nC6PB/j4+BhJtA8OXKz0yEKKFy++cePG7777rnr16i52qlwd41h/wWKx/Pjjj19++WWuxecfFP+A+6lwkyZNateunTkLURTlhx9++OyzzzDrx0Oqh0U+3ErNkiTdvXt38+bNkJfeO6iv7du3Lysry6QMuwtghQMj7SouLg6jMvJUV5gQcujQoYcgyyB/9ff3N7J44CasWrUqT5EIiK8ZGRmYlvi3VKT3gsmxAwAI7amKJa2BNsY1t4YixgX1se789crRC7clKov/fywPNF0AAHDGuUTJ0Qs3t/0WT3x9VU4IMNeMEyCMcau/pUZEMTBz6z+i2p4sy8WLFze/s6gNFArgq3BEFxaCuZADBgz49ddff/zxx+7duxctWhQZCeYAIS8xuc6oiHzyySeJiYmKouRJcKSFuK04xaVLlz755JMmeYXoZh87duyOHTvMm084H1ilSpVCQ0PdBg7hgmfOnKnnWHo+59WrV3tupgOABg0agHGMHed82rRpeTLNEULS09OxNMtDoLw4cyOPK0ZiREdHb9u2jVLqYWNgjPrdvn37tWvX/pZy9A8BHufgYGxfISoWD3gyPBAcDkqNSm4wh6ZMXXFAoNMVVy248JiFCBBcCMYZAExeedChWmXCiHBNV5S4RAkFh6gYbosoHiAEyMjjHicLFtSrV8+Ef6CLWyskwKhXjI93GVFvd2Sz2SIjI1etWhUbG7tx48ZRo0Y1atQoJCQEfR4mtYVQhE1MTFy1ahXksf2lXIgbissICwtbvnx58+bNHQ6HW4OSXnurX79+0dHRZcqUMZd2kRyHhITUqlVr586dOctUYPJaXFzctGnTxo8fjxqZ+VRRQzpw4MDGjRs97GeJa6lVq1a5cuUuXryYk1Ci/XHNmjVRUVEvvviiJ2HKWKpr6dKl586dK0jby7xCq1at5s+fbxQpSAgZPXp0kyZNfH19c81C118ybdo0+MeF2+oRz2lpaUWLFn18F6IxzdeqtKlV+tTZk8QgfoJxkPwsPx+8/O3uuNebV3MwDSSZgESBAVAP3BOcCs3OJKusLNoVu+FAvBQQAiydU9nZe06AMWJViANU1qrek/6ypDHtngPksYLmzZtPnTrVrbSE3un33nsP+60V/FKgGvHMM8+AQfEL52bexYsXb9++PVZrv3Hjxq+//rpt27bNmzefO3fOhLgRQvbu3fv222/nabZy4e4p0uKGDRt+/vnnb775plGXQL3LTZ8+fbZv365HKJsYeSilPXv23LFjh1F8sCRJEyZMeOqpp1588UVsZ29UxAap9p07d9566y3P5UpCiKqqPj4+rVq1WrhwoVHaDgC89tpre/furVq1KvahNApIczgcNpstNjb2ww8/fGhiO+5J69atw8LCEhMTjbhgbGzsoEGDVq5ciduFTjm3246pT5MmTcIef39vU94HZK9ISUm5fv16cHDw484guzeuMGdDLOPgtpyIAIlwDfxDR8/bXb1EUKMqpVRHprD4KkAJeFJglGRqkp8iHTidMGL+QfAtClwFIPeaD4r/14YICCaEYtG6NcB8QHgIhVwLXbBo0KBByZIlExIScl4iSqnD4bDb7ePGjXtABjQjY4Zu0ULDfvHixTt27NixY8eMjIyvvvrqww8/dNtoC5+Pi4vD6+/M8/h9nzoBINR17MI/NlmWVVV944033njjDSQ9RkYPi8Wyd+/ekSNH5hrOiyS4Q4cOISEhbkvT6Ly3R48e33//vdVqxbQal4LJ+FuLxXL9+vWXX375xIkTeSLcOO6AAQOMLFSoCd6+fbtVq1YHDhzAouL60M7ToJTabLZjx4516dIlOTkZHpbbADsPBgUFdenSBYwLeUmStGrVqp49e6anp6MWlXMVqBErijJr1qwJEyZ47kN6vMgEOhVXr17tNuJDrwn/iC+ESjIT4pkKYQ2qhossza0JSwBwARLJTlGtL03etufUDcXiSznjnEFuGX6cC02AnyLtO53w0pTtKQ6FSg4QqkZswClx2h9BZIWoLFttXDW8UcUwLgSlMuonjxFnVlU1MDAwMjLS7SXCG/TFF19gCmFWVhYWDPYcsG6/C5g0WESSgtQMyzihzKc3PvDx8Rk7dmz//v3dVlfCNycmJjoHo2L3F0qIRIlECaWEALhUC3ggbB8Ls3zxxReNGjUySm/RBds5c+YsXrwYuY4J/9A0rXjx4q+++qpJdUysy//qq6/2798fCzGhFoKAf6enpy9btqxx48a7du3Kaxd05EkNGzZs2rSpkT0RP09ISGjVqtX48eNv3LihD+08jRs3bkyfPr1Zs2bnz583yiB9cCwEAN5++22TcXUW0qhRo59//hkX5bIKSmlMTEyPHj1Gjhxp3kD0cbFTmViHZ82aFR0djX0BnJ2TeFcffaWEABGcK7I0pH01UB1ACAUNxF/CaglwIIQzIVuV65lKl482LN75h0wliUoa01TGNGwohRsG7F4XFsYZ55QShZJFO090+mj9zXRNssqcc0GoAOFimyIAnFLg2mvta8qUisfL6fHXSzR48GCbzebWQoUmkw8++ODTTz/18fGxWCzO6XtGTnKdFWGxdxcPvNtRMJlDkqS4uLiYmBhUfXR8xq8URcEPzQvAOByqqmkoSTDuIIRQQk7F31obff6nfWf3n06wg5AoaEzjwAQwXuj2K53cYzHaH374oWHDhkYFzxH/JEkaOnRojRo16tWrZ+IIQb3s3XffXbRoUVpampFnBTW4JUuW/Pjjj61atWrevHnp0qVDQkLS09MTEhJ+//33Xbt2oR0wr8wD7ntiKKUTJ05s1aqV0WMoC9jt9ilTpixcuPCFF15o2LBhyZIl/f3979y5c/Xq1WPHjm3fvh0bbj98mw8u/Kmnnurduzc2kzdpHn7y5Mlu3brVqlWrbdu21atXDw8PF0IkJiaeOXPm4MGDu3btAqfm5I8vYNcHExUkLS2tffv2kyZN6t27NxqyEJKTk0+cOKF3hn9kGYkAkClwzrs1Kv9ltd+PnMtSbDYisjUiEReNilCNc0XW0hgd8PnBn2Muv/9y3QYVS7ioKZRQtJrjdY0+e2PmTzGrD14G36JWOUNlGhAqQOQsTSIBqFm8ceUikY2e5IxhbfnHDvDWVKtWrU+fPt98803OK6B3rBk3blxsbOz06dNLliwJ9zOpnfFErxqOgggSwNTUVPPKoTrxpJRevnx53rx5c+fOlSRp69atDRo0cFGL0dqRkZGxbNkyt6YOrC8TElI0wN8fQIBgErWcS0z96Jt9Px+9nqkBMJlIWQ3KBfzr1RZdapfimp3LVgm4/OD2V9O0iIiIJUuWYKtUtywEtyAzM7N3797R0dHBwcFGDluksxEREWPGjBk/fryRZwU3TpblrKysqKioqKgot3NzW7Xfc7xp2bJlZGTkTz/9ZEQ6dS0yMTFx6dKlS5cudWvoQ//B3yI9CSEmTpy4YcMGrIrvdhp6Icjjx49jXLLRKh5r9wbcT+E20kJQaLhz587bb789ZcqUGjVqFC1aNDs7++bNm+fPn7958+bJkycfcf6Bkr8QzFdWJr5Sr/OEnUT4MMqJO8cDAVCFRCUi+8kb9l/acjSh7dPFuzYs17RaqfCQIF+bJBPCQWRmq9eS7+47eW3d4Wvbjl9Ts4TkHwpcZRwElYxNXgQE+/CVxj6ypDJOgTy+mIOXaP369bdv33YrhyE1W758+c6dO999993evXs/8cQTRiKyEOL69esxMTFRUVE7d+5ctmwZ2jlcnr/X+I9SSZLi4+PnzJnzzTffYHsbAGjTps20adP69evnUmbit99+e++999zWSQMAQighvFr16jabTXNkyxbrqZspncevO3tdQICPjToECJUERp/P6jbh5x9GPd/zuRoOLiTywPiHLpa2bdv2k08+ef/997FVnFsiJcvy2bNnBwwYsG7dOtwvo/5fmqaNHj168+bN+/fvN1EgUKdzLnavN63DajAFXBrn/Isvvti7d+/NmzeNiC+yqJzTwL9R8fzbrOGUappWvnz56dOnv/HGG0ZHo4sqaFrV01P1Veh9aR5fwMtZq1Ytf39/LIvglougQEApTUhIwIRTHaxWa15b2vwdvAMESIRSzkSHOuVfbX1+0ZazcnAAV5nb3AsqOOHACJUCAh1CRMXcjjoU7x8gFwv2L+KnWBXi0HhqunojxZGeqgIlxFe2+HPOMgSAJilEaG7dGZJMtZT0fu0qtasdwblGHnzDtAeKOYyx0qVLz549u0+fPm4NCbo4e+PGjXHjxk2ZMqVBgwbNmjUrU6ZMWFiYoihZWVkpKSmXLl06derU6dOnL168qNcx+ve//71t2zYjzePatWvz58+fP38+lm5CYksIuXv37pAhQ2bNmtW6desaNWr4+fldvHhx3759e/fuVVXVxN0rhHi++XMAwEFyMD503p6z11TfIrKqZdqJDYCBEIqvr+qwDP3PntqVS1UuHsy59GDxHrWQcePGHT9+fNWqVUbSOvpINmzY8PHHH3/00UdGka9ItmRZXrJkSePGjW/evGmkhegn5yEe5EkJ0CsHLF26FIPkTAw4Hk5D7/L7MBVwTdMGDx58+PDh7777DvuvmQjgnmzRAy0e/OCkSM55mTJl6tSps2/fPhNzYk6BAH0h+Ssc9PANWAQAgAgCgqufvtZsz5k7F+NTLVYfhwBCGAATTm04BKECCBDOOaNEyD7AiX864+k3MoALEAIIBUkGSbYESACaBuAQMiFcEIK9cMVfQn4JCJAkomVqFcsGzBjQVAgBBCR4vKO98RK98sor0dHRc+bMwcqwOZEBk+HQCrpjx44dO3aYvxOpwfbt2zdt2tS+fXtnkwxinaqqbdu2PXXqlG4AQPqjG/DPnj2bsxiEkcCNomFISEivnj2ZAItF2RN7Zd/RS1KRItkaFaAIohIhADTGhGJTku9Iy3ee+vcrjVXB6YO+nOikXbBgQfXq1U2SCtGX/vHHH69fv96oQC++jTH25JNPrl69OjAwEIuFFGR6nuf6uUxDVdU2bdrMmzcPj6QgbVgKtyNYXo/mP//5T/v27T1Jmsn1Lj2m/nNEgH79+nligNK1LoxDe2yWTIDcC8EEDWixANuyd5v4KVzjnEogcSZxCn+xZQkg/P6SiSoI44wSIVkUycci+VolH0VSgBJVFUwVRAggggMAEYIILkD6i/4hiIWojEGgnL5oRJuwAB/OGSX0HxCuhxTp888/79atm0nKl7OjO2dPe3SPO/eaRbyaMGEC/srFBuPj49OpUycAwBIezkiox2I5J67rvTaMlsA5/9eYMSVLlFBVOwAcOZekcSsBIggDopJ7PY9BECK4ShR538nrAkAiQB8CnRJCBAYGrlixAut7G5FaNBG89tprZ8+eNbGqI+1u2rTphg0bMFNfUZR82J31emERERH5s86pqjpo0KBvvvlGT4PIx3uwYIDFYilRogQ83PQCHMtqta5evbpTp04mSTO5vgfrCJQoUcKTbmuPIAngnL/yyivVqlUzKb3zzwACQiKSyrXGlUoseK812DOEoITKguQeCCUAuBCc3/9PeJgxLmTKVfCl9rQFw1s1qVhcY0yiAPBPyDUl9+GHH3546aWXHA6HySVyFj74fdDT1J0VWdQ5jh49umrVKhedGNWFDz/8sE6dOij2GcVl4UDmlX3R8NCyZcsRI0ZojMmUA0BatkaAGtVJE1S6k2HXQND70RMP/H5ib/oFCxaYRFjhV8nJyb169crMzDSx56CC8txzz+3du7d27dpoePE8jBJ5PhrN5s2bt2rVKuIOzKk5Ek1N0wYOHLhx48YSJUqgeTFPDXRxZ0JDQ3/55ZeJEyfC/fw+z6ehf5unybvgoo+Pz9q1a0eMGIGoZl4tx2VoZMOoUB86dKhSpUq6hSfnlB7EEgr+HhRxrFbrwoULUaCTHme7fK52LIrBNIz0blp+/nvN5YxklVOhECDag6ihLlGqMU7sd+e926JHk+oOjcsSASL9M/gH3G/QhNXUhw0bpl+igr9ZluUPPvgA41d1YoiKvp+f36pVq8qVK+dwOPInQCMFczgcderUWbFihSzLhICgFgAoWcQmBHMSDoTTjyjRtLLhRRQgjOeIvih4kxMTFtKrV6/Ro0drmoYux5yA3ZyOHj36zjvvmNtD0OVQrVo1zEC0Wq1oTNB1Q5cN1fVE9F4wxho2bLh79+4333yzWrVqpUqVwrGEE+j/zHVdL7zwQnR0dM+ePXXnvNE0dAUWRQzGWLdu3aKjo1u3bt28eXOr1arbMT2fRs6Zezh5ZwcAIWTmzJnr16+vXLmynhuo69Quz+vaMVaBDg4OnjFjxqZNm8qWLVu/fn20AjlPBgWrB7cET97jiRWicePGS5YsQVzKtXypsy3iseIfwIhEBNiIUJn6esvKi8a19RN2nskVKguwEBBUMJGf+u2CCCGA3Pd8CAAqS5KWzfwhe8mYloNaVlMddiJRAfQ+bfqHJJzq/rAvv/zy+++/Dw8Px4tsVLghV8Oy7tVwOBy3bt1y8SyiRlKhQoVdu3bVrVsXczs8H8tZ7Gvfvv3WrVvDwsI4Z5QiXxctapX2DaRcIxJwIoQAiwBKBKdCA8kiNHvH+mXvF1D+6/VLTk7GSefMfszOztY0zcNuGTlZNFL8qVOntm7dGsuBuYXs7GwA+Pbbbz/++GM09pmwEM65v7//Z599dvjw4b59+wYFBem6oe5KQgqo64mc87p16y5atGjfvn1NmjTJzs729/fv1KkTcmP6V8i1E7WuhTzxxBMrVqzYvn17u3btUIZ1mYbuIcdpSJLUqlWrTZs2rVmzpnz58qqqVqxYsU6dOkbTMFefqTvQJ58rVumG106dOsXExMyaNatKlSr6drnspB48pmlaWFjYu+++e/To0VGjRuHnkZGR+tDOFl69soLR6EZL8PBK6Apovt+DKmnPnj337Nnz1FNPuWQIOstPOBDyTh1j83QRjCB/XrQ8vZMCSABACaGSIikqY680qbR5cvvKJRQ1NcsGTCZUEJnmx6+DnnOgghMhKJEppWqqvWq4vPGTLq80qcQ0h2xR5PuemPs/8XQhHvJpc/HXE3Qyr5trNA29R2yfPn0OHz48YMAA3Z2OOGMSVuosi+gkIjw8/P333z9x4kSNGjVyNjdCo2tERMR///vfUaNG6WNh4qFbhNdHQdQNDAz89NNPo6KiQkNDOedUkglQhQBjolKJ4Hc6VufJd4nsJ8kgg2YBVSJEtviqKdnPVA/p3bQ8FyBTKrso/tOnT09OTnZ71XF3sIB5PhBd36Zly5atXr3aXL1Al4DD4bBareY3BxlDrVq1li5devny5XXr1u3cufPkyZNXrlxx9iyFhoZWrFgRe8c3a9YMF6hpGuaFTp48eciQITkxg3OOyWK5NlVGtaNVq1atWrX6448/1q1bt3fv3tOnT8fHxzsL0SVKlKhatWrTpk27du369NNP604txOzVq1cnJye7nQYmHzl/hSRs+fLlWVlZbvESi056eFh4NIhVw4cPHzJkyPbt2zdu3HjkyJELFy6kpKToS7DZbGXLlq1Zs2a7du06duyIrUT0TJHnnnsuNjY2pyMa5SNfX1/9K3y+RIkSR44cMXJc48KNKg4470NAQMC+ffvc1rbx8D3gVL3t8OHDa9euXbRo0a+//pqSkuLWfBwWFlapUqWaNWu2aNGibNmyHrYvS09PR67j8jl+gj3E8goZGRn5fqdCqENzPFulxH+ndx/93aHvt8cBpRZfCxMWkccWtgKIIEQSGqGSoBaWmQk8s3/byp8OaBYe6GPXmCxZCHAAblT2IjMz0+1CELDMjzlwzrF7Zk7A1xr1n3aGu3fvmuyn0fudL1HZsmW/++67YcOGzZkzZ926dc4z182qLhq2jmMWi6Vx48Zdu3bt0aOHfrlMUuL8/PxmzJjRu3fvzz//fM2aNc5ddZ3FVhQEcZTg4OAePXqMGjWqfPnySD9dCAvn/ONejW/ezly05SxYFFAkABk4hdRbtSsFLRnV3s9m04SQCDzsgMsHlGOlpyngP7Ozs69cuZKUlJSWlqYoSpEiRUqXLh0WFqY/X5CW8Sagx3fiP9PS0q5cuXLnzp3MzEyr1Vq0aNHSpUsXKVJE3wqTosp/m30jx6xu3LgRHx9/9+5dxlhAQEB4eHipUqX0YC2XJf8DwNlFl5CQcP78+VOnTt26dQtlmsDAwLCwsIiIiLJlyzpjlCevJYRs2LDh6NGjFovFhSdhtfy6detiXI2HmIlT3bFjx759+3K2QpAkyeFwNGzYsH379m79jgLAAcLKmcYlWQYAbX3MtY9+PPb76ZtAZeojS5RwzoVnBdwlApSAKmSRlQ2ao27VkAk96nWqXxGAc6ZSycKB0HtJhdTtQvbv379p0yZFUVxCIlFyDQ0NHTZsmJFREQlLSkrKV199Zbfb3Urfmqb16NGjVq1aRl5YfMmKFStiY2Nz9ujEaVSqVAmrKJngvJ7iBwDXrl3bunXrrl27Dh06dO3atZwMTFGUkJCQcuXK1a5du379+g0bNqxSpYrnZMr5wp47d27Tpk3btm377bffbt265YwPhJBixYo1atSoTZs2HTt2LFOmjOn7BYCqgWX1ntjFO84cS8hknJcOkV5uUHrIiw2K+tq4YIRQAjkC9nMNSUQeW0AK5WFKRF4H0msDGFWm0rPAckrHRqv23ITiMhBaTtzyGMQtN/c579MwDz7Ox+SdkdJZhHEri7m9zCZTMipFXChLKMSt0FM9zDHQHN/MOVNhCVgexhwbPcOEABAUmABJ49wiyXbNsWr/+Tkb446cT4ZsDawWsEgSlYgAAI4eCyHudSm858QghDEu7CqoHGxyk/K+Q16o2a1ZVR9Z1jgjBAjBZTMASRTAbY67bV6A9iEIuObTcEEP3ZGOQu2NGzdSU1OxBLivr6/NZgsODi5ZsqRzY2zUCfLkO9EL7uoq1KVLl5KSktLT01VV9fPzK1asWEREhF56R09JMdwHLgAYoTIA3EzLEBoEB/tYgQIwximlQIBxkMk/r2aqswfVRWf835xGoeh2j/sqCqKNudwRD4PKTGiKCX/NhyZXoHcKIQjBBoACCONCpgBAGPB9J6/9uP/Cvrjb568mZ2VrIAAoAUkCSpGJAAfgDAQBImRfqFwy5LlqxXs8W65JtVISEADBOGCc7v3+H8K8xq75Qjzh07nKpp5EaZrL0HmVaz0UMvSZ5w8HPJFc9Qc8ClUVAAQY5xK5110YQHDGyf8LlAKA/DP5hxe84IV8skwAzoUEAigFgGyNx12+/evFxAsJyZdvZ964k5GW6dAYs1Di7+cbHuxbOiyoQomAZ8qHVitT1CZL9yghJ5IkEe9uGkuTOcWRBzpWQUa59xp3TWC8/MMLXvCCK2gAjAtFaFQiLlW6OaDJi+RQJpiDCQGSQggVArzs438AvPzDC17wgluZUwiighBCKJq4H5xLqASY3XG/N7oAAoISAkQAcCpkIIQ/nMxkL3j5hxe84IVHDzgACKCAtQ/veS7+0goKO9Hed24Q9K1L974SBLz6h5d/eMELXvCCF7zgDrxaphe84AUveMHLP7zgBS94wQsPC/4PyGe/BI9R4TcAAAAASUVORK5CYII="

# ─── Helpers ───────────────────────────────────────────────────────────────────

def month_key(dt):
    if not isinstance(dt, datetime):
        return None
    return f"{dt.year}-{str(dt.month).zfill(2)}"


def month_label(key):
    months = ["", "January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    y, m = key.split("-")
    return f"{months[int(m)]} {y}"


def month_short(key):
    months = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    y, m = key.split("-")
    return f"{months[int(m)]} '{y[2:]}"


def days_in_month(key):
    y, m = key.split("-")
    return calendar.monthrange(int(y), int(m))[1]


def next_month_label(key):
    months = ["", "January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    y, m = int(key.split("-")[0]), int(key.split("-")[1])
    nm = m + 1 if m < 12 else 1
    ny = y if m < 12 else y + 1
    return f"{months[nm]} {ny}"


def fmt(n, decimals=2):
    return f"{n:,.{decimals}f}"


def format_date(dt):
    if not isinstance(dt, datetime):
        return ""
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{dt.day} {months[dt.month - 1]}"


# ─── Data Extraction ──────────────────────────────────────────────────────────

def load_unit_registry(wb):
    ws = wb["Unit Registry"]
    units = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        code = row[1].value
        building = row[2].value
        model = row[3].value
        if not code or not building or str(model).strip() != "Revenue Share":
            continue
        if str(code).strip() == "Unit Code":
            continue
        units.append({
            "code": str(code).strip(),
            "building": str(building).strip(),
            "owner": str(row[6].value or "[Owner Name]").strip(),
            "email": str(row[7].value or "[Email]").strip(),
            "phone": str(row[8].value or "[Phone]").strip(),
            "active": str(row[10].value or "").strip(),
        })
    return units


def get_available_months(wb, units):
    """Get available months from the Sales sheet (works with formula-based P&L)."""
    months_set = set()
    unit_codes = {u["code"] for u in units}
    ws = wb["Sales"]
    header_row = None
    for r in range(1, 6):
        val = ws.cell(row=r, column=1).value
        if val and "Hostaway" in str(val):
            header_row = r
            break
    if header_row is None:
        return []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        prop = str(row[3].value or "").strip()
        sm = row[7].value
        if prop in unit_codes and isinstance(sm, datetime):
            months_set.add(month_key(sm))
    return sorted(months_set, reverse=True)


def safe_float(value, default=0.0):
    """Safely convert a cell value to float, returning default on failure."""
    try:
        return float(value or default)
    except (ValueError, TypeError):
        return default


def load_expenses(wb, unit_code, month):
    """Compute owner expenses from Expenses sheet using P&L SUMIFS logic."""
    ws = wb["Expenses"]
    y, m = month.split("-")
    target_year, target_month = int(y), int(m)
    utilities = 0.0
    reimbursement = 0.0
    for r in range(4, ws.max_row + 1):
        cc = str(ws.cell(row=r, column=15).value or "").strip()
        if cc != unit_code:
            continue
        subcat = str(ws.cell(row=r, column=16).value or "").strip()
        ac_cat = str(ws.cell(row=r, column=17).value or "").strip()
        svc_month = ws.cell(row=r, column=3).value
        exp_date = ws.cell(row=r, column=1).value
        amount = safe_float(ws.cell(row=r, column=6).value)
        if (subcat == "Utilities" and
            ac_cat not in ("Apartment Startup Cost", "Business Startup Cost") and
            isinstance(svc_month, datetime) and
            svc_month.year == target_year and svc_month.month == target_month):
            utilities += amount
        if (ac_cat == "Apartment Startup Cost" and
            isinstance(exp_date, datetime) and
            exp_date.year == target_year and exp_date.month == target_month):
            reimbursement += amount
        if (subcat == "Reimbursement" and
            isinstance(exp_date, datetime) and
            exp_date.year == target_year and exp_date.month == target_month):
            reimbursement += amount
    return round(utilities, 2), round(reimbursement, 2)


def load_pnl(wb, unit_code, month, bookings=None):
    """Compute P&L from raw Sales + Expenses data (no formula dependency)."""
    if bookings is None:
        bookings = load_bookings(wb, unit_code, month)
    if not bookings:
        return None

    utilities, reimbursement = load_expenses(wb, unit_code, month)

    total_gross = round(sum(b["guest_paid"] for b in bookings), 2)
    platform_fees = round(sum(b["host_fee_total"] for b in bookings), 2)
    payment_charges = round(sum(b["payment_charges"] for b in bookings), 2)
    net_earned = round(sum(b["remitted"] for b in bookings), 2)
    cleaning_retained = round(sum(b["cleaning"] for b in bookings), 2)
    tourism_retained = round(sum(b["tourism"] for b in bookings), 2)
    rev_net_retained = round(net_earned - cleaning_retained - tourism_retained, 2)
    total_owner_expenses = round(utilities + reimbursement, 2)
    net_before_mgmt = round(rev_net_retained - total_owner_expenses, 2)
    mgmt_fee = round(rev_net_retained * 0.15, 2)
    owner_payout = round(net_before_mgmt - mgmt_fee, 2)

    if total_gross == 0:
        return None

    return {
        "total_gross": total_gross, "platform_fees": platform_fees,
        "payment_charges": payment_charges, "net_earned": net_earned,
        "cleaning_retained": -cleaning_retained, "tourism_retained": -tourism_retained,
        "rev_net_retained": rev_net_retained,
        "utilities": utilities, "reimbursement": reimbursement,
        "total_owner_expenses": -total_owner_expenses,
        "net_before_mgmt": net_before_mgmt,
        "mgmt_fee": -mgmt_fee, "owner_payout": owner_payout,
    }


def load_bookings(wb, unit_code, month):
    """Extract bookings from Sales, computing formula columns from raw data."""
    ws = wb["Sales"]
    header_row = None
    for r in range(1, 6):
        val = ws.cell(row=r, column=1).value
        if val and "Hostaway" in str(val):
            header_row = r
            break
    if header_row is None:
        return []

    bookings = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        prop = row[3].value
        sale_month = row[7].value
        if (str(prop).strip() != unit_code or
            not isinstance(sale_month, datetime) or
            month_key(sale_month) != month):
            continue

        # Raw input columns (M through V + AK)
        m_to_v = sum(safe_float(row[c].value) for c in range(12, 22))
        ak = safe_float(row[36].value) if len(row) > 36 else 0

        # Computed: GuestPaid = SUM(M:V) + AK
        guest_paid = round(m_to_v + ak, 2)

        # Direct values
        host_fee_total = safe_float(row[25].value)   # Z
        pg_fees_total = safe_float(row[28].value)    # AC
        refunds = safe_float(row[29].value)          # AD
        other_receipt = safe_float(row[30].value)    # AE

        # Computed: Remitted = GuestPaid + HostFee + PGFees - Refunds + OtherReceipt
        remitted = round(guest_paid + host_fee_total + pg_fees_total - refunds + other_receipt, 2)

        bookings.append({
            "guest": str(row[2].value or ""),
            "platform": str(row[6].value or ""),
            "checkin": row[8].value,
            "checkout": row[9].value,
            "nights": int(row[10].value or 0),
            "cleaning": round(safe_float(row[13].value), 2),
            "tourism": round(safe_float(row[15].value), 2),
            "guest_paid": guest_paid,
            "host_fee_total": host_fee_total,
            "payment_charges": pg_fees_total,
            "remitted": remitted,
        })
    return bookings


# ─── SOA Calculation ──────────────────────────────────────────────────────────

def calculate_soa(unit, pnl, bookings, month):
    running_pm = 0.0
    rows = []

    for i, b in enumerate(bookings):
        rev = b["guest_paid"]
        cleaning = b["cleaning"]
        tourism = b["tourism"]
        commission = round(abs(b["host_fee_total"]) + abs(b["payment_charges"]), 2)
        net = b["remitted"]
        rev_net_ret = net - cleaning - tourism
        pm = round(rev_net_ret * 0.15, 2)
        running_pm += pm
        gross = round(rev_net_ret - pm, 2)

        ch_class, ch_label = "ch-direct", "Direct"
        plat = b["platform"].lower()
        if "airbnb" in plat:
            ch_class, ch_label = "ch-airbnb", "Airbnb"
        elif "booking" in plat:
            ch_class, ch_label = "ch-booking", "Booking"

        rows.append({
            "num": i + 1, "guest": b["guest"], "ch_class": ch_class, "ch_label": ch_label,
            "checkin": format_date(b["checkin"]), "checkout": format_date(b["checkout"]),
            "nights": b["nights"], "rev": rev, "cleaning": cleaning, "commission": commission,
            "net": net, "pm": pm, "gross": gross,
        })

    # Adjust PM to match PnL
    pm_target = abs(pnl["mgmt_fee"])
    pm_diff = round(running_pm - pm_target, 2)
    if abs(pm_diff) > 0.001 and rows:
        largest = max(rows, key=lambda r: r["net"])
        largest["pm"] = round(largest["pm"] - pm_diff, 2)
        largest["gross"] = round(largest["gross"] + pm_diff, 2)

    totals = {k: round(sum(r[k] for r in rows), 2) for k in ["rev", "cleaning", "commission", "net", "pm", "gross"]}
    totals["nights"] = sum(r["nights"] for r in rows)

    available = days_in_month(month)
    utilities = pnl.get("utilities", 0)
    reimbursement = pnl.get("reimbursement", 0)
    expenses = abs(pnl["total_owner_expenses"])
    fees_received = round(abs(pnl["cleaning_retained"]) + abs(pnl["tourism_retained"]), 2)
    total_ded = round(fees_received + expenses + abs(pnl["mgmt_fee"]) + abs(pnl["platform_fees"]) + abs(pnl["payment_charges"]), 2)

    unit_number = unit["code"].split(" ")[1] if " " in unit["code"] else unit["code"]

    return {
        "unit": unit, "month": month,
        "property_name": f"{unit['building']} {unit_number}",
        "rows": rows, "totals": totals, "available": available,
        "expenses": expenses, "fees_received": fees_received,
        "deductions": {
            "fees_received": fees_received, "utilities": utilities,
            "reimbursement": reimbursement,
            "mgmt_fee": abs(pnl["mgmt_fee"]), "platform_fees": abs(pnl["platform_fees"]),
            "payment_charges": abs(pnl["payment_charges"]), "total": total_ded,
        },
        "kpi": {
            "owner_gross": round(totals["gross"]),
            "occupancy": round((totals["nights"] / available) * 100),
            "booked": totals["nights"], "available": available,
            "reservations": len(rows),
            "net_payout": round(pnl["owner_payout"]),
            "net_payout_exact": round(pnl["owner_payout"], 2),
        },
    }


# ─── HTML Template ────────────────────────────────────────────────────────────

def generate_html(soa, logo_b64=None):
    u = soa["unit"]
    k = soa["kpi"]
    t = soa["totals"]
    d = soa["deductions"]
    m = soa["month"]
    y, mn = m.split("-")
    ms = ["", "January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
    period = f"{ms[int(mn)]} 1 — {days_in_month(m)}, {y}"

    actual_logo = logo_b64 if logo_b64 else RADIANT_LOGO_B64
    logo = f'<img src="data:image/png;base64,{actual_logo}" style="height:36px;width:auto">'

    reimb_row = (
        f'<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px">'
        f'<span style="color:#6b7280">Reimbursement Cost</span>'
        f'<span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["reimbursement"])}</span></div>'
        if d["reimbursement"] else ""
    )

    brows = ""
    for r in soa["rows"]:
        dim = ' style="color:#6b7280;opacity:0.45"' if r["rev"] == 0 else ""
        brows += f'''<tr><td{dim}>{r["num"]}</td><td{dim}>{r["guest"]}</td>
        <td><span class="ch {r["ch_class"]}">{r["ch_label"]}</span></td>
        <td{dim}>{r["checkin"]}</td><td{dim}>{r["checkout"]}</td>
        <td class="r"{dim}>{r["nights"]}</td><td class="r"{dim}>{fmt(r["rev"])}</td><td class="r"{dim}>{fmt(r["cleaning"])}</td>
        <td class="r"{dim}>{fmt(r["commission"])}</td><td class="r"{dim}>{fmt(r["net"])}</td><td class="r"{dim}>{fmt(r["pm"])}</td><td class="r"{dim}>{fmt(r["gross"])}</td></tr>'''

    return f'''<!DOCTYPE html><html><head><meta charset="UTF-8">
<link href="https://fonts.bunny.net/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Outfit',sans-serif;background:#fff;color:#1a1d24;-webkit-font-smoothing:antialiased}}
.page{{width:210mm;margin:0 auto;position:relative;overflow:hidden}}
.page::before{{content:'';position:absolute;top:0;left:0;right:0;height:5px;background:linear-gradient(90deg,#0d4a75,#1565a0,#0d4a75)}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
thead th{{text-align:left;padding:10px 6px;font-size:8.5px;letter-spacing:1.2px;text-transform:uppercase;color:#6b7280;background:#f7f9fb;border-bottom:2px solid #1565a0;font-weight:600;white-space:nowrap}}
thead th.r,tbody td.r,tfoot td.r{{text-align:right}}
tbody td{{padding:9px 6px;border-bottom:1px solid #eaeff4;white-space:nowrap;font-variant-numeric:tabular-nums}}
tfoot td{{padding:12px 6px;font-weight:700;border-top:2px solid #1565a0;background:#f7f9fb;font-size:11.5px}}
.ch{{display:inline-block;font-size:8px;font-weight:700;letter-spacing:.8px;text-transform:uppercase;padding:3px 8px;border-radius:4px}}
.ch-airbnb{{background:rgba(217,79,79,.07);color:#d94f4f}}.ch-booking{{background:rgba(21,101,160,.07);color:#1565a0}}.ch-direct{{background:rgba(26,138,106,.07);color:#1a8a6a}}
</style></head><body>
<div class="page">
<div style="padding:32px 44px 0;display:flex;justify-content:space-between;align-items:center;position:relative;z-index:1">{logo}<div style="font-size:10px;font-weight:600;letter-spacing:2.5px;text-transform:uppercase;color:#1565a0;border:1.5px solid #1565a0;padding:6px 16px;border-radius:6px">Owner's Statement</div></div>
<div style="padding:26px 44px 22px;display:flex;justify-content:space-between;align-items:flex-end;border-bottom:1px solid #dde3ea"><div><div style="font-size:36px;font-weight:700;letter-spacing:-.5px;line-height:1.1">{soa["property_name"]}</div><div style="font-size:13px;color:#6b7280;margin-top:6px">{period}</div></div><div style="text-align:right"><div style="font-size:9px;letter-spacing:2.5px;text-transform:uppercase;color:#6b7280;margin-bottom:6px">Net Owner Payout</div><div style="font-size:42px;font-weight:800;color:#1565a0;letter-spacing:-1px"><span style="font-size:16px;font-weight:400;color:#6b7280">AED </span>{k["net_payout"]:,}</div></div></div>
<div style="padding:14px 44px;background:#f7f9fb;border-bottom:1px solid #eaeff4;display:flex;gap:40px;font-size:12px;color:#6b7280"><span><strong style="color:#1a1d24;font-weight:500">Owner:</strong> {u["owner"]}</span><span><strong style="color:#1a1d24;font-weight:500">Phone:</strong> {u["phone"]}</span><span><strong style="color:#1a1d24;font-weight:500">Email:</strong> {u["email"]}</span></div>
<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:22px 44px">
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:20px;text-align:center;position:relative"><div style="position:absolute;top:0;left:28%;right:28%;height:3px;border-radius:0 0 3px 3px;background:#1565a0"></div><div style="font-size:26px;font-weight:700;color:#1565a0;margin-bottom:4px;letter-spacing:-.5px">{k["owner_gross"]:,}</div><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#6b7280;font-weight:500">Owner Gross (AED)</div></div>
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:20px;text-align:center;position:relative"><div style="position:absolute;top:0;left:28%;right:28%;height:3px;border-radius:0 0 3px 3px;background:#1a8a6a"></div><div style="font-size:26px;font-weight:700;color:#1a8a6a;margin-bottom:4px;letter-spacing:-.5px">{k["occupancy"]}%</div><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#6b7280;font-weight:500">Occupancy</div></div>
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:20px;text-align:center;position:relative"><div style="position:absolute;top:0;left:28%;right:28%;height:3px;border-radius:0 0 3px 3px;background:#c08b2e"></div><div style="font-size:26px;font-weight:700;color:#c08b2e;margin-bottom:4px;letter-spacing:-.5px">{k["booked"]} / {k["available"]}</div><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#6b7280;font-weight:500">Booked / Available</div></div>
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:20px;text-align:center;position:relative"><div style="position:absolute;top:0;left:28%;right:28%;height:3px;border-radius:0 0 3px 3px;background:#d94f4f"></div><div style="font-size:26px;font-weight:700;color:#d94f4f;margin-bottom:4px;letter-spacing:-.5px">{k["reservations"]}</div><div style="font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:#6b7280;font-weight:500">Reservations</div></div>
</div>
<div style="font-size:9px;font-weight:700;letter-spacing:3px;text-transform:uppercase;color:#1565a0;padding:14px 44px 10px;display:flex;align-items:center;gap:14px">Rental Activity Details — {ms[int(mn)]} {y}<span style="flex:1;height:1px;background:#dde3ea"></span></div>
<div style="padding:0 44px 12px"><table><thead><tr><th>#</th><th>Guest</th><th>Channel</th><th>In</th><th>Out</th><th class="r">Nts</th><th class="r">Booking Rev</th><th class="r">Cleaning</th><th class="r">Commission</th><th class="r">Net Rev</th><th class="r">PM 15%</th><th class="r">Gross</th></tr></thead><tbody>{brows}</tbody>
<tfoot><tr><td colspan="5">Total</td><td class="r">{t["nights"]}</td><td class="r">{fmt(t["rev"])}</td><td class="r">{fmt(t["cleaning"])}</td><td class="r">{fmt(t["commission"])}</td><td class="r">{fmt(t["net"])}</td><td class="r">{fmt(t["pm"])}</td><td class="r">{fmt(t["gross"])}</td></tr></tfoot></table></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;padding:0 44px 18px">
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:22px"><div style="font-size:9px;letter-spacing:2.5px;text-transform:uppercase;color:#1565a0;font-weight:700;margin-bottom:14px;padding-bottom:10px;border-bottom:1px solid #dde3ea">Expenses & Extras</div><div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Utilities & Service Charge — {month_short(m)}</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["utilities"])}</span></div>{reimb_row}<div style="display:flex;justify-content:space-between;padding:10px 0 6px;font-size:12px;font-weight:700;border-top:1.5px solid #1a1d24;margin-top:8px"><span>Total Expenses</span><span style="color:#d94f4f;font-variant-numeric:tabular-nums">AED {fmt(soa["expenses"])}</span></div></div>
<div style="border:1.5px solid #dde3ea;border-radius:12px;padding:22px"><div style="font-size:9px;letter-spacing:2.5px;text-transform:uppercase;color:#1565a0;font-weight:700;margin-bottom:14px;padding-bottom:10px;border-bottom:1px solid #dde3ea">Deductions Breakdown</div>
<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Fees Received (Cleaning + Tourism)</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["fees_received"])}</span></div>
<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Utilities & Service Charge</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["utilities"])}</span></div>{reimb_row}
<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Management Fee (15%)</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["mgmt_fee"])}</span></div>
<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Platform Host Fees</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["platform_fees"])}</span></div>
<div style="display:flex;justify-content:space-between;padding:6px 0;font-size:12px"><span style="color:#6b7280">Payment Charges</span><span style="font-weight:500;font-variant-numeric:tabular-nums">{fmt(d["payment_charges"])}</span></div>
<div style="display:flex;justify-content:space-between;padding:10px 0 6px;font-size:12px;font-weight:700;border-top:1.5px solid #1a1d24;margin-top:8px"><span>Total Deductions</span><span style="color:#d94f4f;font-variant-numeric:tabular-nums">AED {fmt(d["total"])}</span></div>
<div style="display:flex;justify-content:space-between;padding:14px 0 0;margin-top:12px;border-top:2px solid #1565a0"><span style="font-size:9px;letter-spacing:2px;text-transform:uppercase;color:#6b7280;align-self:center">Net Owner Payout</span><span style="font-size:22px;font-weight:800;color:#1565a0;font-variant-numeric:tabular-nums">AED {fmt(k["net_payout_exact"])}</span></div></div>
</div>
<div style="margin:0 44px 18px;padding:12px 20px;background:rgba(21,101,160,.07);border-left:3px solid #1565a0;border-radius:0 8px 8px 0;font-size:12px;color:#0d4a75"><strong>Payment Schedule:</strong> {month_short(m)} payout will be processed on <strong>28th {next_month_label(m)}</strong></div>
<div style="padding:12px 44px 20px">
<div style="font-size:9px;font-weight:700;letter-spacing:3px;text-transform:uppercase;color:#1565a0;padding:0 0 12px;display:flex;align-items:center;gap:14px">Notes & Definitions<span style="flex:1;height:1px;background:#dde3ea"></span></div>
<div style="display:flex;gap:10px;margin-bottom:7px;font-size:11px;line-height:1.6;color:#6b7280"><span style="color:#1565a0;font-weight:700;min-width:16px">1.</span><div><strong style="color:#1a1d24;font-weight:500">Booking Revenue:</strong> Total amount collected from the guest including accommodation, cleaning, tourism, VAT, and other fees.</div></div>
<div style="display:flex;gap:10px;margin-bottom:7px;font-size:11px;line-height:1.6;color:#6b7280"><span style="color:#1565a0;font-weight:700;min-width:16px">2.</span><div><strong style="color:#1a1d24;font-weight:500">Commission:</strong> Platform host fees (Airbnb, Booking.com) and payment processing charges.</div></div>
<div style="display:flex;gap:10px;margin-bottom:7px;font-size:11px;line-height:1.6;color:#6b7280"><span style="color:#1565a0;font-weight:700;min-width:16px">3.</span><div><strong style="color:#1a1d24;font-weight:500">Net Revenue:</strong> Amount remitted to Radiant Homes after platform and payment deductions.</div></div>
<div style="display:flex;gap:10px;margin-bottom:7px;font-size:11px;line-height:1.6;color:#6b7280"><span style="color:#1565a0;font-weight:700;min-width:16px">4.</span><div><strong style="color:#1a1d24;font-weight:500">PM 15%:</strong> Property Management Commission calculated at 15% of revenue net of retained fees.</div></div>
<div style="display:flex;gap:10px;margin-bottom:7px;font-size:11px;line-height:1.6;color:#6b7280"><span style="color:#1565a0;font-weight:700;min-width:16px">5.</span><div><strong style="color:#1a1d24;font-weight:500">Owner Gross:</strong> Amount before operational expenses. Owner Gross less Expenses equals Net Owner Payout.</div></div>
</div>
<div style="padding:18px 44px;border-top:1px solid #dde3ea;display:flex;justify-content:space-between;font-size:10px;color:#6b7280"><span>Radiant Vacation Homes Rental L.L.C</span><span>3503, Aspect Tower, Business Bay, UAE</span></div>
</div></body></html>'''


# ─── PDF Generation ───────────────────────────────────────────────────────────

def html_to_pdf(html_content):
    """Convert HTML string to PDF bytes using Playwright."""
    from playwright.sync_api import sync_playwright
    
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False, mode="w", encoding="utf-8") as f:
        f.write(html_content)
        tmp_html = f.name

    tmp_pdf = tmp_html.replace(".html", ".pdf")

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                args=['--no-sandbox', '--disable-setuid-sandbox', '--disable-dev-shm-usage']
            )
            page = browser.new_page()
            page.goto(f"file://{os.path.abspath(tmp_html)}", wait_until="networkidle")
            try:
                page.wait_for_function(
                    "() => document.fonts.ready.then(() => true)",
                    timeout=8000,
                )
            except Exception:
                pass
            page.wait_for_timeout(1000)

            height = page.evaluate("() => document.querySelector('.page') ? document.querySelector('.page').scrollHeight : 1200")
            width = page.evaluate("() => document.querySelector('.page') ? document.querySelector('.page').offsetWidth : 794")

            page.pdf(
                path=tmp_pdf,
                width=f"{width}px",
                height=f"{height + 20}px",
                print_background=True,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )
            browser.close()

        with open(tmp_pdf, "rb") as f:
            return f.read()
    finally:
        for p in [tmp_html, tmp_pdf]:
            if os.path.exists(p):
                try:
                    os.unlink(p)
                except Exception:
                    pass


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/upload", methods=["POST"])
def upload():
    """Parse workbook and return available units + months."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        if not file.filename.endswith((".xlsx", ".xls")):
            return jsonify({"error": "Please upload an .xlsx file"}), 400

        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        units = load_unit_registry(wb)
        months = get_available_months(wb, units)

        # Save workbook to temp file
        file.seek(0)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        file.save(tmp.name)
        tmp.close()

        return jsonify({
            "wb_path": tmp.name,
            "units": units,
            "months": months,
        })
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500


@app.route("/api/generate", methods=["POST"])
def generate():
    """Generate PDFs and return as a zip."""
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No data received"}), 400

        wb_path = data.get("wb_path")
        month = data.get("month")
        unit_codes = data.get("units", [])
        logo_b64 = data.get("logo_b64")

        if not wb_path or not os.path.exists(wb_path):
            return jsonify({"error": "Workbook expired. Please re-upload the file."}), 400

        if not month:
            return jsonify({"error": "No month selected"}), 400

        if not unit_codes:
            return jsonify({"error": "No units selected"}), 400

        wb = openpyxl.load_workbook(wb_path, data_only=True)
        units = load_unit_registry(wb)

        zip_buffer = io.BytesIO()
        results = []

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for code in unit_codes:
                try:
                    unit = next((u for u in units if u["code"] == code), None)
                    if not unit:
                        results.append({"code": code, "status": "error", "msg": "Unit not found"})
                        continue

                    bookings = load_bookings(wb, code, month)
                    if not bookings:
                        results.append({"code": code, "status": "skip", "msg": "No bookings"})
                        continue

                    pnl = load_pnl(wb, code, month, bookings)
                    if not pnl:
                        results.append({"code": code, "status": "skip", "msg": "No P&L data"})
                        continue

                    soa = calculate_soa(unit, pnl, bookings, month)
                    html = generate_html(soa, logo_b64)
                    pdf_bytes = html_to_pdf(html)

                    unit_num = code.split(" ")[1] if " " in code else code
                    filename = f"{unit['building'].replace(' ', '_')}_{unit_num}_SOA_{month_label(month).replace(' ', '_')}.pdf"
                    zf.writestr(filename, pdf_bytes)

                    results.append({
                        "code": code, "status": "ok",
                        "name": soa["property_name"],
                        "payout": soa["kpi"]["net_payout_exact"],
                        "gross": soa["kpi"]["owner_gross"],
                        "bookings": soa["kpi"]["reservations"],
                        "filename": filename,
                    })
                except Exception as e:
                    results.append({"code": code, "status": "error", "msg": str(e)})

        zip_buffer.seek(0)

        # Save zip
        tmp_zip = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp_zip.write(zip_buffer.getvalue())
        tmp_zip.close()

        return jsonify({"results": results, "zip_path": tmp_zip.name})

    except Exception as e:
        return jsonify({"error": f"Generation failed: {str(e)}\n{traceback.format_exc()}"}), 500


@app.route("/api/download")
def download():
    """Download the generated zip file."""
    zip_path = request.args.get("path")
    if not zip_path or not os.path.exists(zip_path):
        return jsonify({"error": "File not found. Please regenerate."}), 404

    month = request.args.get("month", "statements")
    return send_file(
        zip_path,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"Radiant_Homes_SOAs_{month.replace(' ', '_')}.zip",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Starting SOA Generator on port {port}...")
    app.run(host="0.0.0.0", port=port, debug=False)
